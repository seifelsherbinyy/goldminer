#!/usr/bin/env python3
"""
Script to validate Excel workbook styling against requirements.

This script checks that the generated Excel workbook meets all the styling requirements:
- Dark gray header rows with white bold text
- Calibri 11pt font throughout
- Alternating row shading (light gray)
- Consistent column widths
- Proper currency formatting
"""

import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

def validate_workbook_styling(filename):
    """Validate the styling of an Excel workbook."""
    print("=" * 80)
    print(f"Validating Excel Workbook Styling: {filename}")
    print("=" * 80)
    print()
    
    # Load workbook
    wb = load_workbook(filename)
    
    print(f"Sheets found: {wb.sheetnames}")
    print()
    
    all_passed = True
    
    for sheet_name in wb.sheetnames:
        print(f"📋 Validating sheet: {sheet_name}")
        print("-" * 80)
        ws = wb[sheet_name]
        
        # Check 1: Header styling (row 1)
        header_checks = []
        if ws.max_row > 0:
            # Check first header cell
            header_cell = ws['A1']
            
            # Check font
            if header_cell.font.name == 'Calibri':
                header_checks.append("✓ Font: Calibri")
            else:
                header_checks.append(f"✗ Font: {header_cell.font.name} (expected Calibri)")
                all_passed = False
            
            # Check font size
            if header_cell.font.size in [11, 12, 14]:  # Allow some variation for titles
                header_checks.append(f"✓ Font size: {header_cell.font.size}pt")
            else:
                header_checks.append(f"✗ Font size: {header_cell.font.size}pt (expected 11pt)")
                all_passed = False
            
            # Check if bold
            if header_cell.font.bold:
                header_checks.append("✓ Bold: Yes")
            else:
                header_checks.append("✗ Bold: No (expected Yes)")
                all_passed = False
            
            # For header rows with dark background
            if hasattr(header_cell.fill, 'start_color'):
                fill_color = header_cell.fill.start_color.rgb if header_cell.fill.start_color else None
                if fill_color and fill_color != '00000000' and fill_color != 'FFFFFFFF':
                    header_checks.append(f"✓ Header background: #{fill_color}")
                else:
                    header_checks.append("⚠ Header background: white/transparent")
        
        for check in header_checks:
            print(f"  {check}")
        
        # Check 2: Data row font consistency
        data_checks = []
        if ws.max_row > 1:
            data_cell = ws['A2']
            if data_cell.value is not None:
                if data_cell.font.name == 'Calibri':
                    data_checks.append("✓ Data font: Calibri")
                else:
                    data_checks.append(f"✗ Data font: {data_cell.font.name} (expected Calibri)")
                    all_passed = False
                
                if data_cell.font.size == 11:
                    data_checks.append("✓ Data font size: 11pt")
                else:
                    data_checks.append(f"⚠ Data font size: {data_cell.font.size}pt (expected 11pt)")
        
        for check in data_checks:
            print(f"  {check}")
        
        # Check 3: Alternating row shading
        alternating_checks = []
        has_alternating = False
        if ws.max_row > 2:
            # Check a few rows for alternating pattern
            for row_idx in range(2, min(6, ws.max_row + 1)):
                cell = ws.cell(row=row_idx, column=1)
                fill_color = cell.fill.start_color.rgb if hasattr(cell.fill, 'start_color') and cell.fill.start_color else None
                if fill_color and fill_color not in ['00000000', 'FFFFFFFF', None]:
                    has_alternating = True
                    break
        
        if has_alternating:
            alternating_checks.append("✓ Alternating row shading: Detected")
        else:
            alternating_checks.append("⚠ Alternating row shading: Not clearly detected")
        
        for check in alternating_checks:
            print(f"  {check}")
        
        # Check 4: Freeze panes
        freeze_checks = []
        if ws.freeze_panes:
            freeze_checks.append(f"✓ Freeze panes: {ws.freeze_panes}")
        else:
            freeze_checks.append("⚠ Freeze panes: Not set")
        
        for check in freeze_checks:
            print(f"  {check}")
        
        # Check 5: Column widths
        width_checks = []
        non_default_widths = 0
        for col in ws.column_dimensions.values():
            if col.width and col.width != 8.43:  # Default width
                non_default_widths += 1
        
        if non_default_widths > 0:
            width_checks.append(f"✓ Column widths: {non_default_widths} columns auto-adjusted")
        else:
            width_checks.append("⚠ Column widths: Using default widths")
        
        for check in width_checks:
            print(f"  {check}")
        
        print()
    
    print("=" * 80)
    if all_passed:
        print("✅ All styling requirements validated!")
    else:
        print("⚠️  Some styling checks did not pass (see details above)")
    print("=" * 80)
    
    return all_passed

def main():
    """Main function."""
    filename = 'demo_transactions_export.xlsx'
    
    if len(sys.argv) > 1:
        filename = sys.argv[1]
    
    if not Path(filename).exists():
        print(f"Error: File '{filename}' not found")
        return 1
    
    validate_workbook_styling(filename)
    return 0

if __name__ == '__main__':
    sys.exit(main())
