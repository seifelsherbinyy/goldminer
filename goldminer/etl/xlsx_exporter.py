"""
XLSX Exporter module for exporting transaction data to Excel workbooks.

This module provides the XLSXExporter class for creating well-formatted Excel workbooks
with multiple sheets, charts, and formatting suitable for non-technical users.
"""
from typing import List, Dict, Optional, Any
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import PieChart, BarChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import CellIsRule
from goldminer.utils import setup_logger


class XLSXExporter:
    """
    Exports transaction data to well-formatted Excel workbooks.
    
    Creates multi-sheet workbooks with:
    - Transactions: Full row-level data
    - Monthly Summary: Aggregated totals by category/account
    - Anomalies: Records with anomaly flags
    
    Features:
    - Formatted headers with freeze panes
    - Auto-adjusted column widths
    - Currency formatting for amounts
    - Highlighted anomaly rows
    - Charts: pie (category share), bar (monthly spend), line (cumulative)
    - UTF-8 compatible
    """
    
    def __init__(self):
        """Initialize the XLSX exporter."""
        self.logger = setup_logger(__name__)
        
        # Define reusable style templates for professional appearance
        # Font: Calibri 11pt (sans-serif) for all cells
        self.default_font = Font(name="Calibri", size=11)
        
        # Header style: Dark gray background with white bold text
        self.header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        self.header_fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Alternating row shading for data tables (light gray)
        self.alternating_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Border style for all cells
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Currency/numeric cell style
        self.currency_font = Font(name="Calibri", size=11)
        self.currency_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        
        # Tagged text style (for anomaly tags)
        self.tag_font = Font(name="Calibri", size=11, italic=True, color="D9534F")
        self.anomaly_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Define urgency-based formatting styles
        self.urgency_high_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        self.urgency_high_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        self.urgency_medium_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        self.urgency_low_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        # Define anomaly border style
        self.anomaly_border = Border(
            left=Side(style='thick', color="FF0000"),
            right=Side(style='thick', color="FF0000"),
            top=Side(style='thick', color="FF0000"),
            bottom=Side(style='thick', color="FF0000")
        )
        
        self.logger.info("XLSXExporter initialized")
    
    def _apply_urgency_formatting(self, ws, urgency_col_letter: str, start_row: int, end_row: int) -> None:
        """
        Apply conditional formatting for urgency levels.
        
        Uses openpyxl's conditional formatting rules to dynamically format cells
        based on their urgency value without hardcoding row indices.
        
        Args:
            ws: Worksheet object
            urgency_col_letter: Column letter containing urgency values (e.g., 'J')
            start_row: First data row (usually 2, after header)
            end_row: Last data row
        """
        if end_row < start_row:
            return
        
        # Define the range for the entire row based on urgency column
        urgency_range = f"{urgency_col_letter}{start_row}:{urgency_col_letter}{end_row}"
        
        # High urgency: red fill with bold white text
        high_rule = CellIsRule(
            operator='equal',
            formula=['"high"'],
            fill=self.urgency_high_fill,
            font=self.urgency_high_font
        )
        ws.conditional_formatting.add(urgency_range, high_rule)
        
        # Medium urgency: yellow fill
        medium_rule = CellIsRule(
            operator='equal',
            formula=['"medium"'],
            fill=self.urgency_medium_fill
        )
        ws.conditional_formatting.add(urgency_range, medium_rule)
        
        # Low/Normal urgency: green tint
        low_rule = CellIsRule(
            operator='equal',
            formula=['"normal"'],
            fill=self.urgency_low_fill
        )
        ws.conditional_formatting.add(urgency_range, low_rule)
        
        # Also apply for "low" in case it appears in data
        low_alt_rule = CellIsRule(
            operator='equal',
            formula=['"low"'],
            fill=self.urgency_low_fill
        )
        ws.conditional_formatting.add(urgency_range, low_alt_rule)
        
        self.logger.info(f"Applied urgency conditional formatting to column {urgency_col_letter}")
    
    def _apply_row_urgency_formatting(self, ws, urgency_col_idx: int, start_row: int, end_row: int, 
                                       row_start_col: int, row_end_col: int) -> None:
        """
        Apply conditional formatting to entire rows based on urgency in a specific column.
        
        This extends formatting across all columns in a row when urgency is detected.
        
        Args:
            ws: Worksheet object
            urgency_col_idx: Column index containing urgency values
            start_row: First data row
            end_row: Last data row
            row_start_col: First column to format in each row
            row_end_col: Last column to format in each row
        """
        if end_row < start_row:
            return
        
        from openpyxl.utils import get_column_letter
        
        urgency_col_letter = get_column_letter(urgency_col_idx)
        
        # Apply formatting to all columns in rows with specific urgency
        for col_idx in range(row_start_col, row_end_col + 1):
            col_letter = get_column_letter(col_idx)
            cell_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
            
            # High urgency: red fill with bold white text
            high_rule = CellIsRule(
                operator='equal',
                formula=[f'${urgency_col_letter}{start_row}="high"'],
                fill=self.urgency_high_fill,
                font=self.urgency_high_font
            )
            ws.conditional_formatting.add(cell_range, high_rule)
            
            # Medium urgency: yellow fill
            medium_rule = CellIsRule(
                operator='equal',
                formula=[f'${urgency_col_letter}{start_row}="medium"'],
                fill=self.urgency_medium_fill
            )
            ws.conditional_formatting.add(cell_range, medium_rule)
            
            # Low/Normal urgency: green tint
            low_rule = CellIsRule(
                operator='equal',
                formula=[f'${urgency_col_letter}{start_row}="normal"'],
                fill=self.urgency_low_fill
            )
            ws.conditional_formatting.add(cell_range, low_rule)
            
            # Also handle "low" value
            low_alt_rule = CellIsRule(
                operator='equal',
                formula=[f'${urgency_col_letter}{start_row}="low"'],
                fill=self.urgency_low_fill
            )
            ws.conditional_formatting.add(cell_range, low_alt_rule)
        
        self.logger.info(f"Applied row-based urgency formatting based on column {urgency_col_letter}")
    
    def _apply_anomaly_borders(self, ws, anomaly_col_idx: int, start_row: int, end_row: int,
                                row_start_col: int, row_end_col: int) -> None:
        """
        Apply bold red borders to rows containing anomaly flags.
        
        Checks for non-empty anomaly values and applies border formatting to the entire row.
        
        Args:
            ws: Worksheet object
            anomaly_col_idx: Column index containing anomaly flags
            start_row: First data row
            end_row: Last data row
            row_start_col: First column to format in each row
            row_end_col: Last column to format in each row
        """
        if end_row < start_row:
            return
        
        from openpyxl.utils import get_column_letter
        
        anomaly_col_letter = get_column_letter(anomaly_col_idx)
        
        # Apply borders to all columns in rows with anomalies
        for col_idx in range(row_start_col, row_end_col + 1):
            col_letter = get_column_letter(col_idx)
            cell_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
            
            # Create rule for non-empty anomaly values
            anomaly_rule = CellIsRule(
                operator='notEqual',
                formula=['""'],
                border=self.anomaly_border
            )
            # Apply with reference to anomaly column
            for row_idx in range(start_row, end_row + 1):
                anomaly_value = ws.cell(row=row_idx, column=anomaly_col_idx).value
                if anomaly_value and str(anomaly_value).strip():
                    # Apply border to each cell in the row
                    for col in range(row_start_col, row_end_col + 1):
                        ws.cell(row=row_idx, column=col).border = self.anomaly_border
        
        self.logger.info(f"Applied anomaly borders to rows with flags in column {anomaly_col_letter}")
    
    def _apply_alternating_row_shading(self, ws, start_row: int, end_row: int, 
                                        start_col: int, end_col: int) -> None:
        """
        Apply alternating light gray shading to data rows for better scannability.
        
        This creates a professional, easy-to-read table with zebra striping.
        
        Args:
            ws: Worksheet object
            start_row: First data row (usually 2, after header)
            end_row: Last data row
            start_col: First column to shade
            end_col: Last column to shade
        """
        if end_row < start_row:
            return
        
        # Apply light gray fill to every other row
        for row_idx in range(start_row, end_row + 1):
            # Apply shading to even rows (0-indexed, so row 2, 4, 6, etc.)
            if (row_idx - start_row) % 2 == 1:
                for col_idx in range(start_col, end_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    # Only apply if cell doesn't already have a fill (preserve urgency formatting)
                    if cell.fill.patternType is None or cell.fill.start_color.rgb in ['00000000', 'FFFFFFFF']:
                        cell.fill = self.alternating_fill
        
        self.logger.info(f"Applied alternating row shading to rows {start_row}-{end_row}")
    
    def export_to_excel(self, transactions: List[Dict[str, Any]], filename: str) -> None:
        """
        Export transactions to a well-formatted Excel workbook.
        
        Args:
            transactions: List of transaction dictionaries with keys like:
                - date, amount, payee, category, subcategory, account_id, 
                  account_type, anomalies, etc.
            filename: Output filename for the Excel workbook
            
        Raises:
            ValueError: If transactions list is empty or filename is invalid
            IOError: If unable to write to the specified file
        """
        if not transactions:
            raise ValueError("Cannot export empty transaction list")
        
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        self.logger.info(f"Exporting {len(transactions)} transactions to {filename}")
        
        # Convert to DataFrame for easier manipulation
        df = pd.DataFrame(transactions)
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Create sheets
        self._create_transactions_sheet(wb, df)
        self._create_monthly_summary_sheet(wb, df)
        self._create_anomalies_sheet(wb, df)
        
        # Save workbook
        try:
            wb.save(filename)
            self.logger.info(f"Successfully exported to {filename}")
        except Exception as e:
            self.logger.error(f"Failed to save workbook: {e}")
            raise IOError(f"Unable to write to file {filename}: {e}")
    
    def _create_transactions_sheet(self, wb: Workbook, df: pd.DataFrame) -> None:
        """
        Create the Transactions sheet with full row-level data.
        
        Args:
            wb: Workbook object
            df: DataFrame with transaction data
        """
        ws = wb.create_sheet("Transactions", 0)
        
        # Define column order and headers
        columns = [
            'id', 'date', 'payee', 'category', 'subcategory', 
            'amount', 'currency', 'account_id', 'account_type',
            'urgency', 'tags', 'anomalies', 'confidence'
        ]
        
        # Filter columns that exist in the dataframe
        available_columns = [col for col in columns if col in df.columns]
        df_subset = df[available_columns].copy()
        
        # Write data to sheet
        for r_idx, row in enumerate(dataframe_to_rows(df_subset, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Style header row
                if r_idx == 1:
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = self.header_alignment
                    cell.border = self.border
                else:
                    # Apply default font and border to data cells
                    cell.font = self.default_font
                    cell.border = self.border
                    
                    # Apply currency formatting to amount column
                    if available_columns[c_idx - 1] == 'amount' and value is not None:
                        cell.number_format = self.currency_format
                        cell.font = self.currency_font
        
        # Apply alternating row shading (before conditional formatting)
        self._apply_alternating_row_shading(
            ws,
            start_row=2,
            end_row=len(df_subset) + 1,
            start_col=1,
            end_col=len(available_columns)
        )
        
        # Apply conditional formatting for urgency (if urgency column exists)
        if 'urgency' in available_columns:
            urgency_col_idx = available_columns.index('urgency') + 1
            from openpyxl.utils import get_column_letter
            urgency_col_letter = get_column_letter(urgency_col_idx)
            
            # Apply row-based urgency formatting (formats entire row based on urgency value)
            self._apply_row_urgency_formatting(
                ws, 
                urgency_col_idx, 
                start_row=2, 
                end_row=len(df_subset) + 1,
                row_start_col=1,
                row_end_col=len(available_columns)
            )
        
        # Apply anomaly borders (if anomalies column exists)
        if 'anomalies' in available_columns:
            anomaly_col_idx = available_columns.index('anomalies') + 1
            self._apply_anomaly_borders(
                ws,
                anomaly_col_idx,
                start_row=2,
                end_row=len(df_subset) + 1,
                row_start_col=1,
                row_end_col=len(available_columns)
            )
        
        # Freeze top row
        ws.freeze_panes = ws['A2']
        
        # Auto-adjust column widths
        self._auto_adjust_column_widths(ws)
        
        self.logger.info(f"Created Transactions sheet with {len(df)} rows")
    
    def _create_monthly_summary_sheet(self, wb: Workbook, df: pd.DataFrame) -> None:
        """
        Create the Monthly Summary sheet with dashboard-style layout.
        
        Layout:
        - Top-left: High-level metrics (total spend, # of transactions, credit vs debit split)
        - Top-right: Charts area (line, bar, pie)
        - Bottom: Detailed monthly pivot table (category × month × amount)
        
        Args:
            wb: Workbook object
            df: DataFrame with transaction data
        """
        ws = wb.create_sheet("Monthly Summary", 1)
        
        # Ensure date column exists and is properly formatted
        if 'date' not in df.columns or 'amount' not in df.columns:
            self.logger.warning("Missing required columns for monthly summary")
            return
        
        # Convert date to datetime if it's not already
        df_copy = df.copy()
        df_copy['date'] = pd.to_datetime(df_copy['date'], errors='coerce')
        df_copy = df_copy.dropna(subset=['date'])
        
        if df_copy.empty:
            self.logger.warning("No valid dates found for monthly summary")
            return
        
        # Extract year-month
        df_copy['year_month'] = df_copy['date'].dt.to_period('M').astype(str)
        
        # Get currency symbol from data if available
        currency_symbol = 'EGP' if 'currency' not in df_copy.columns else df_copy['currency'].iloc[0] if len(df_copy) > 0 else 'EGP'
        
        # ===== DASHBOARD HEADER (Row 1) =====
        ws['A1'] = 'GoldMiner Financial Dashboard'
        ws['A1'].font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells('A1:N1')
        ws.row_dimensions[1].height = 30
        
        # ===== TOP-LEFT: HIGH-LEVEL METRICS (Rows 3-11) =====
        metrics_title_font = Font(name="Calibri", bold=True, size=12, color="1F4E78")
        metrics_label_font = Font(name="Calibri", size=11)
        metrics_value_font = Font(name="Calibri", bold=True, size=14)
        metrics_fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
        
        # Section header
        ws['A3'] = 'Key Metrics'
        ws['A3'].font = Font(name="Calibri", bold=True, size=13, color="1F4E78")
        ws.merge_cells('A3:E3')
        ws['A3'].alignment = Alignment(horizontal="left", vertical="center")
        
        # Calculate metrics
        total_spend = df_copy['amount'].sum()
        num_transactions = len(df_copy)
        
        # Credit vs Debit split
        credit_amount = 0
        debit_amount = 0
        if 'account_type' in df_copy.columns:
            credit_amount = df_copy[df_copy['account_type'].str.lower() == 'credit']['amount'].sum() if any(df_copy['account_type'].str.lower() == 'credit') else 0
            debit_amount = df_copy[df_copy['account_type'].str.lower() == 'debit']['amount'].sum() if any(df_copy['account_type'].str.lower() == 'debit') else 0
        
        # Metric 1: Total Spend
        ws['A5'] = '  Total Spend:'
        ws['B5'] = total_spend
        ws['A5'].font = metrics_label_font
        ws['B5'].font = metrics_value_font
        ws['B5'].number_format = f'"{currency_symbol} "#,##0.00'
        ws['B5'].fill = metrics_fill
        ws['B5'].alignment = Alignment(horizontal="right", vertical="center", indent=1)
        ws.merge_cells('B5:E5')
        
        # Metric 2: Number of Transactions
        ws['A7'] = '  # of Transactions:'
        ws['B7'] = num_transactions
        ws['A7'].font = metrics_label_font
        ws['B7'].font = metrics_value_font
        ws['B7'].number_format = '#,##0'
        ws['B7'].fill = metrics_fill
        ws['B7'].alignment = Alignment(horizontal="right", vertical="center", indent=1)
        ws.merge_cells('B7:E7')
        
        # Metric 3: Credit Split
        ws['A9'] = '  Credit Card:'
        ws['B9'] = credit_amount
        ws['A9'].font = metrics_label_font
        ws['B9'].font = metrics_value_font
        ws['B9'].number_format = f'"{currency_symbol} "#,##0.00'
        ws['B9'].fill = metrics_fill
        ws['B9'].alignment = Alignment(horizontal="right", vertical="center", indent=1)
        ws.merge_cells('B9:E9')
        
        # Metric 4: Debit Split
        ws['A10'] = '  Debit Card:'
        ws['B10'] = debit_amount
        ws['A10'].font = metrics_label_font
        ws['B10'].font = metrics_value_font
        ws['B10'].number_format = f'"{currency_symbol} "#,##0.00'
        ws['B10'].fill = metrics_fill
        ws['B10'].alignment = Alignment(horizontal="right", vertical="center", indent=1)
        ws.merge_cells('B10:E10')
        
        # Add borders to metrics section
        for row in range(5, 11):
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = Border(
                    left=Side(style='thin', color='B4C7E7'),
                    right=Side(style='thin', color='B4C7E7'),
                    top=Side(style='thin', color='B4C7E7'),
                    bottom=Side(style='thin', color='B4C7E7')
                )
        
        # ===== TOP-RIGHT: CHARTS AREA (Starting from column G, rows 3-20) =====
        # Charts will be added by _add_charts method
        
        # ===== BOTTOM: DETAILED MONTHLY PIVOT TABLE (Starting from row 22) =====
        pivot_start_row = 22
        
        # Pivot section header
        ws[f'A{pivot_start_row}'] = 'Monthly Category Breakdown'
        ws[f'A{pivot_start_row}'].font = Font(name="Calibri", bold=True, size=13, color="1F4E78")
        ws.merge_cells(f'A{pivot_start_row}:N{pivot_start_row}')
        ws[f'A{pivot_start_row}'].alignment = Alignment(horizontal="left", vertical="center")
        
        # Create category × month pivot table
        if 'category' in df_copy.columns:
            category_summary = df_copy.groupby(['year_month', 'category'])['amount'].sum().reset_index()
            category_pivot = category_summary.pivot(index='category', columns='year_month', values='amount').fillna(0)
            
            # Add a total column
            category_pivot['Total'] = category_pivot.sum(axis=1)
            
            # Add a total row
            category_pivot.loc['Total'] = category_pivot.sum(axis=0)
            
            # Write pivot table
            pivot_data_start = pivot_start_row + 2
            category_pivot_reset = category_pivot.reset_index()
            
            # Write headers
            for c_idx, col_name in enumerate(category_pivot_reset.columns, 1):
                cell = ws.cell(row=pivot_data_start, column=c_idx, value=col_name)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.header_alignment
                cell.border = self.border
            
            # Write data rows
            for r_idx, row in enumerate(category_pivot_reset.itertuples(index=False), pivot_data_start + 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.font = self.default_font
                    cell.border = self.border
                    
                    # Format amount columns (all except first column which is category)
                    if c_idx > 1:
                        cell.number_format = f'"{currency_symbol} "#,##0.00'
                        cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                    
                    # Bold the Total row
                    if r_idx == pivot_data_start + len(category_pivot_reset) - 1:
                        cell.font = Font(name="Calibri", bold=True, size=11)
                        cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            
            # Apply alternating row shading (excluding total row)
            self._apply_alternating_row_shading(
                ws,
                start_row=pivot_data_start + 1,
                end_row=pivot_data_start + len(category_pivot_reset) - 2,
                start_col=1,
                end_col=len(category_pivot_reset.columns)
            )
        
        # ===== CONFIGURE PAGE SETUP FOR PRINTING =====
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0  # Allow multiple pages vertically if needed
        
        # ===== FREEZE PANES (Freeze dashboard header and metrics) =====
        ws.freeze_panes = 'A4'
        
        # ===== ADD FOOTER WITH GENERATION DATE =====
        generation_date = datetime.now().strftime('%Y-%m-%d')
        ws.oddFooter.center.text = f"GoldMiner Report — Generated on {generation_date}"
        ws.oddFooter.center.font = "Calibri,Italic"
        ws.oddFooter.center.size = 10
        
        # ===== AUTO-ADJUST COLUMN WIDTHS =====
        self._auto_adjust_column_widths(ws)
        
        # Set specific widths for metrics columns
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        
        # ===== ADD CHARTS =====
        # Overall monthly totals for charts
        monthly_totals = df_copy.groupby('year_month')['amount'].agg(['sum', 'mean', 'count']).reset_index()
        monthly_totals.columns = ['Month', 'Total', 'Average', 'Count']
        self._add_dashboard_charts(ws, monthly_totals, df_copy, currency_symbol)
        
        self.logger.info("Created Monthly Summary sheet with dashboard layout")
    
    def _create_anomalies_sheet(self, wb: Workbook, df: pd.DataFrame) -> None:
        """
        Create the Anomalies sheet with flagged transactions only.
        
        Args:
            wb: Workbook object
            df: DataFrame with transaction data
        """
        ws = wb.create_sheet("Anomalies", 2)
        
        # Filter transactions with anomalies
        if 'anomalies' not in df.columns:
            ws['A1'] = 'No anomaly data available'
            self.logger.warning("No anomalies column found")
            return
        
        # Filter for rows with non-empty anomalies
        df_anomalies = df[df['anomalies'].notna() & (df['anomalies'] != '')].copy()
        
        if df_anomalies.empty:
            ws['A1'] = 'No anomalies detected'
            self.logger.info("No anomalies found")
            return
        
        # Define columns to show
        columns = [
            'id', 'date', 'payee', 'category', 'amount', 
            'account_id', 'urgency', 'anomalies', 'confidence'
        ]
        available_columns = [col for col in columns if col in df_anomalies.columns]
        df_subset = df_anomalies[available_columns].copy()
        
        # Write data to sheet
        for r_idx, row in enumerate(dataframe_to_rows(df_subset, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Style header row
                if r_idx == 1:
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = self.header_alignment
                    cell.border = self.border
                else:
                    # Apply default font and border to data cells
                    cell.font = self.default_font
                    cell.border = self.border
                    
                    # Apply currency formatting to amount column
                    if available_columns[c_idx - 1] == 'amount' and value is not None:
                        cell.number_format = self.currency_format
                        cell.font = self.currency_font
        
        # Apply alternating row shading (before conditional formatting)
        self._apply_alternating_row_shading(
            ws,
            start_row=2,
            end_row=len(df_subset) + 1,
            start_col=1,
            end_col=len(available_columns)
        )
        
        # Apply conditional formatting for urgency (if urgency column exists)
        if 'urgency' in available_columns:
            urgency_col_idx = available_columns.index('urgency') + 1
            from openpyxl.utils import get_column_letter
            urgency_col_letter = get_column_letter(urgency_col_idx)
            
            # Apply row-based urgency formatting
            self._apply_row_urgency_formatting(
                ws, 
                urgency_col_idx, 
                start_row=2, 
                end_row=len(df_subset) + 1,
                row_start_col=1,
                row_end_col=len(available_columns)
            )
        
        # Apply anomaly borders (anomalies column always exists in this sheet)
        if 'anomalies' in available_columns:
            anomaly_col_idx = available_columns.index('anomalies') + 1
            self._apply_anomaly_borders(
                ws,
                anomaly_col_idx,
                start_row=2,
                end_row=len(df_subset) + 1,
                row_start_col=1,
                row_end_col=len(available_columns)
            )
        
        # Freeze top row
        ws.freeze_panes = ws['A2']
        
        # Auto-adjust column widths
        self._auto_adjust_column_widths(ws)
        
        self.logger.info(f"Created Anomalies sheet with {len(df_anomalies)} rows")
    
    def _add_charts(self, ws, monthly_totals: pd.DataFrame, df: pd.DataFrame) -> None:
        """
        Add charts to the Monthly Summary sheet.
        
        Args:
            ws: Worksheet object
            monthly_totals: DataFrame with monthly summary data
            df: Original DataFrame with transaction data
        """
        if monthly_totals.empty:
            return
        
        # Calculate chart positions (to the right of the data)
        chart_col_start = 10
        
        # 1. Pie Chart - Category Share
        if 'category' in df.columns:
            category_totals = df.groupby('category')['amount'].sum().reset_index()
            
            # Write category data for chart
            chart_data_row = 3
            ws.cell(row=chart_data_row, column=chart_col_start, value='Category')
            ws.cell(row=chart_data_row, column=chart_col_start + 1, value='Total')
            for idx, row in enumerate(category_totals.itertuples(), chart_data_row + 1):
                ws.cell(row=idx, column=chart_col_start, value=row.category)
                ws.cell(row=idx, column=chart_col_start + 1, value=row.amount)
            
            # Create pie chart
            pie = PieChart()
            pie.title = "Spending by Category"
            pie.style = 10
            pie.width = 15
            pie.height = 10
            
            # Add data to chart
            data = Reference(ws, min_col=chart_col_start + 1, min_row=chart_data_row, 
                           max_row=chart_data_row + len(category_totals))
            cats = Reference(ws, min_col=chart_col_start, min_row=chart_data_row + 1, 
                           max_row=chart_data_row + len(category_totals))
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(cats)
            
            ws.add_chart(pie, f"J3")
        
        # 2. Bar Chart - Monthly Spend
        bar = BarChart()
        bar.type = "col"
        bar.title = "Monthly Spending"
        bar.y_axis.title = 'Amount'
        bar.x_axis.title = 'Month'
        bar.width = 15
        bar.height = 10
        
        data_rows = len(monthly_totals) + 3
        data = Reference(ws, min_col=2, min_row=3, max_row=data_rows, max_col=2)
        cats = Reference(ws, min_col=1, min_row=4, max_row=data_rows)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        
        ws.add_chart(bar, "J20")
        
        # 3. Line Chart - Cumulative Spending
        line = LineChart()
        line.title = "Cumulative Spending Trend"
        line.y_axis.title = 'Cumulative Amount'
        line.x_axis.title = 'Month'
        line.width = 15
        line.height = 10
        
        # Calculate cumulative sum
        cumulative_col = 5
        ws.cell(row=3, column=cumulative_col, value='Cumulative')
        cumulative_sum = 0
        for idx, row in enumerate(monthly_totals.itertuples(), 4):
            cumulative_sum += row.Total
            ws.cell(row=idx, column=cumulative_col, value=cumulative_sum)
        
        data = Reference(ws, min_col=cumulative_col, min_row=3, max_row=data_rows)
        cats = Reference(ws, min_col=1, min_row=4, max_row=data_rows)
        line.add_data(data, titles_from_data=True)
        line.set_categories(cats)
        
        ws.add_chart(line, "J37")
        
        self.logger.info("Added charts to Monthly Summary sheet")
    
    def _add_dashboard_charts(self, ws, monthly_totals: pd.DataFrame, df: pd.DataFrame, currency_symbol: str = 'EGP') -> None:
        """
        Add charts to the dashboard layout in the top-right area.
        
        Args:
            ws: Worksheet object
            monthly_totals: DataFrame with monthly summary data
            df: Original DataFrame with transaction data
            currency_symbol: Currency symbol to use in charts
        """
        if monthly_totals.empty:
            return
        
        # Chart positions in top-right area (starting from column G, row 3)
        chart_col_start = 7  # Column G
        
        # 1. Pie Chart - Category Share (Top Right)
        if 'category' in df.columns:
            category_totals = df.groupby('category')['amount'].sum().reset_index()
            category_totals = category_totals.sort_values('amount', ascending=False)
            
            # Write category data for chart in a hidden area
            hidden_data_col = 16  # Column P (far right, won't interfere)
            chart_data_row = 3
            ws.cell(row=chart_data_row, column=hidden_data_col, value='Category')
            ws.cell(row=chart_data_row, column=hidden_data_col + 1, value='Total')
            for idx, row in enumerate(category_totals.itertuples(), chart_data_row + 1):
                ws.cell(row=idx, column=hidden_data_col, value=row.category)
                ws.cell(row=idx, column=hidden_data_col + 1, value=row.amount)
            
            # Create pie chart
            pie = PieChart()
            pie.title = "Spending by Category"
            pie.style = 10
            pie.width = 12
            pie.height = 9
            
            # Add data to chart
            data = Reference(ws, min_col=hidden_data_col + 1, min_row=chart_data_row, 
                           max_row=chart_data_row + len(category_totals))
            cats = Reference(ws, min_col=hidden_data_col, min_row=chart_data_row + 1, 
                           max_row=chart_data_row + len(category_totals))
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(cats)
            
            # Position in top-right
            ws.add_chart(pie, "G3")
        
        # 2. Bar Chart - Monthly Spending Trends (Middle Right)
        # Write monthly data for bar chart
        hidden_data_col = 18  # Column R
        bar_data_row = 3
        ws.cell(row=bar_data_row, column=hidden_data_col, value='Month')
        ws.cell(row=bar_data_row, column=hidden_data_col + 1, value='Spending')
        for idx, row in enumerate(monthly_totals.itertuples(), bar_data_row + 1):
            ws.cell(row=idx, column=hidden_data_col, value=row.Month)
            ws.cell(row=idx, column=hidden_data_col + 1, value=row.Total)
        
        bar = BarChart()
        bar.type = "col"
        bar.title = "Monthly Spending"
        bar.y_axis.title = f'Amount ({currency_symbol})'
        bar.x_axis.title = 'Month'
        bar.width = 12
        bar.height = 7
        
        data_rows = bar_data_row + len(monthly_totals)
        data = Reference(ws, min_col=hidden_data_col + 1, min_row=bar_data_row, max_row=data_rows)
        cats = Reference(ws, min_col=hidden_data_col, min_row=bar_data_row + 1, max_row=data_rows)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        
        # Position below pie chart
        ws.add_chart(bar, "G13")
        
        # 3. Line Chart - Cumulative Spending Trend (Bottom Right)
        # Calculate and write cumulative data
        hidden_data_col = 20  # Column T
        line_data_row = 3
        ws.cell(row=line_data_row, column=hidden_data_col, value='Month')
        ws.cell(row=line_data_row, column=hidden_data_col + 1, value='Cumulative')
        cumulative_sum = 0
        for idx, row in enumerate(monthly_totals.itertuples(), line_data_row + 1):
            cumulative_sum += row.Total
            ws.cell(row=idx, column=hidden_data_col, value=row.Month)
            ws.cell(row=idx, column=hidden_data_col + 1, value=cumulative_sum)
        
        line = LineChart()
        line.title = "Cumulative Spending"
        line.y_axis.title = f'Cumulative ({currency_symbol})'
        line.x_axis.title = 'Month'
        line.width = 6
        line.height = 7
        line.style = 13
        
        data = Reference(ws, min_col=hidden_data_col + 1, min_row=line_data_row, max_row=line_data_row + len(monthly_totals))
        cats = Reference(ws, min_col=hidden_data_col, min_row=line_data_row + 1, max_row=line_data_row + len(monthly_totals))
        line.add_data(data, titles_from_data=True)
        line.set_categories(cats)
        
        # Position to the right of bar chart
        ws.add_chart(line, "K13")
        
        self.logger.info("Added dashboard charts to Monthly Summary sheet")
    
    def _auto_adjust_column_widths(self, ws) -> None:
        """
        Auto-adjust column widths based on content.
        
        Args:
            ws: Worksheet object
        """
        from openpyxl.cell.cell import MergedCell
        
        for column in ws.columns:
            max_length = 0
            column_letter = None
            
            for cell in column:
                # Skip merged cells
                if isinstance(cell, MergedCell):
                    continue
                    
                if column_letter is None:
                    column_letter = cell.column_letter
                
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # Only adjust if we found a valid column letter
            if column_letter:
                # Set width with some padding, but cap at reasonable maximum
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = max(adjusted_width, 10)
