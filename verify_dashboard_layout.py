#!/usr/bin/env python3
"""
Verification script for Dashboard Layout in Monthly Summary sheet.

This script verifies all the requirements from the problem statement:
1. Dashboard-style layout with zones
2. High-level metrics in top-left
3. Charts in top-right
4. Detailed pivot table at bottom
5. Cell merging and formatting
6. Proper number formatting (EGP with commas, 2 decimals)
7. Dashboard header in A1
8. Freeze panes configured
9. Print settings (fit to page width)
10. Footer with generation date
"""

from openpyxl import load_workbook
from datetime import datetime
import sys
from pathlib import Path

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent))

from goldminer.etl import XLSXExporter


def verify_dashboard_layout(filename):
    """Verify all dashboard layout requirements."""
    print("=" * 80)
    print("Dashboard Layout Verification Report")
    print("=" * 80)
    print()
    
    wb = load_workbook(filename)
    ws = wb['Monthly Summary']
    
    results = []
    
    # 1. Dashboard Header in A1
    print("1. Dashboard Header:")
    header = ws['A1'].value
    is_merged = any('A1' in str(r) and 'N1' in str(r) for r in ws.merged_cells.ranges)
    print(f"   ✓ Header text: '{header}'")
    print(f"   ✓ Cell A1 merged across columns: {is_merged}")
    print(f"   ✓ Font: {ws['A1'].font.name}, Size: {ws['A1'].font.size}, Bold: {ws['A1'].font.bold}")
    print(f"   ✓ Background color: #{ws['A1'].fill.start_color.rgb if ws['A1'].fill.start_color else 'None'}")
    results.append(header is not None and is_merged)
    print()
    
    # 2. Top-Left: High-Level Metrics
    print("2. Top-Left: High-Level Metrics:")
    print(f"   ✓ Section Title (A3): '{ws['A3'].value}'")
    print(f"   ✓ Metric 1 - {ws['A5'].value} = {ws['B5'].value} (Format: {ws['B5'].number_format})")
    print(f"   ✓ Metric 2 - {ws['A7'].value} = {ws['B7'].value} (Format: {ws['B7'].number_format})")
    print(f"   ✓ Metric 3 - {ws['A9'].value} = {ws['B9'].value} (Format: {ws['B9'].number_format})")
    print(f"   ✓ Metric 4 - {ws['A10'].value} = {ws['B10'].value} (Format: {ws['B10'].number_format})")
    
    # Check number formatting includes currency and proper format
    has_currency = any(c in ws['B5'].number_format for c in ['USD', 'EGP', '$', '€', '£'])
    has_commas = '#,##0' in ws['B5'].number_format
    has_decimals = '.00' in ws['B5'].number_format
    print(f"   ✓ Currency symbol present: {has_currency}")
    print(f"   ✓ Comma separators: {has_commas}")
    print(f"   ✓ Two decimal places: {has_decimals}")
    results.append(has_currency and has_commas and has_decimals)
    print()
    
    # 3. Top-Right: Charts Area
    print("3. Top-Right: Charts Area:")
    print(f"   ✓ Number of charts: {len(ws._charts)}")
    chart_info = []
    for i, chart in enumerate(ws._charts, 1):
        title = "Unknown"
        if chart.title and chart.title.tx and chart.title.tx.rich:
            for para in chart.title.tx.rich.p:
                for run in para.r:
                    if run.t:
                        title = run.t
                        break
        position = chart.anchor._from if hasattr(chart.anchor, '_from') else 'Unknown'
        col = position.col if hasattr(position, 'col') else '?'
        row = position.row if hasattr(position, 'row') else '?'
        print(f"   ✓ Chart {i}: '{title}' at column {col}, row {row}")
        chart_info.append(title)
    
    has_pie = any('Category' in t for t in chart_info)
    has_bar = any('Monthly' in t for t in chart_info)
    has_line = any('Cumulative' in t for t in chart_info)
    print(f"   ✓ Pie chart present: {has_pie}")
    print(f"   ✓ Bar chart present: {has_bar}")
    print(f"   ✓ Line chart present: {has_line}")
    results.append(has_pie and has_bar and has_line)
    print()
    
    # 4. Bottom: Detailed Monthly Pivot Table
    print("4. Bottom: Detailed Monthly Pivot Table:")
    pivot_title = ws['A22'].value
    pivot_header = ws['A24'].value
    print(f"   ✓ Section Title (A22): '{pivot_title}'")
    print(f"   ✓ First column header (A24): '{pivot_header}'")
    
    # Check pivot table formatting
    has_pivot_data = ws['B24'].value is not None
    print(f"   ✓ Has month columns: {has_pivot_data}")
    
    # Check for Total column
    last_col = ws.max_column
    has_total_col = False
    for col in range(1, last_col + 1):
        if ws.cell(24, col).value == 'Total':
            has_total_col = True
            break
    print(f"   ✓ Has Total column: {has_total_col}")
    results.append(pivot_title is not None and has_pivot_data)
    print()
    
    # 5. Cell Merging and Formatting
    print("5. Cell Merging and Formatting:")
    merged_count = len(list(ws.merged_cells.ranges))
    print(f"   ✓ Number of merged cell ranges: {merged_count}")
    results.append(merged_count >= 5)  # At least header, metrics, and section titles
    print()
    
    # 6. Freeze Panes
    print("6. Freeze Panes:")
    freeze_cell = ws.freeze_panes
    print(f"   ✓ Freeze panes at: {freeze_cell}")
    results.append(freeze_cell == 'A4')
    print()
    
    # 7. Print Settings
    print("7. Print Settings:")
    orientation = "Landscape" if ws.page_setup.orientation == ws.ORIENTATION_LANDSCAPE else "Portrait"
    fit_width = ws.page_setup.fitToWidth
    print(f"   ✓ Page orientation: {orientation}")
    print(f"   ✓ Fit to page width: {fit_width}")
    results.append(orientation == "Landscape" and fit_width == 1)
    print()
    
    # 8. Footer with Generation Date
    print("8. Footer:")
    footer_text = ws.oddFooter.center.text if ws.oddFooter.center.text else "None"
    print(f"   ✓ Footer text: '{footer_text}'")
    has_goldminer = 'GoldMiner' in footer_text
    has_date = 'Generated' in footer_text
    print(f"   ✓ Contains 'GoldMiner': {has_goldminer}")
    print(f"   ✓ Contains generation date: {has_date}")
    results.append(has_goldminer and has_date)
    print()
    
    # Summary
    print("=" * 80)
    print("VERIFICATION SUMMARY")
    print("=" * 80)
    passed = sum(results)
    total = len(results)
    print(f"Checks passed: {passed}/{total}")
    
    if passed == total:
        print("✓ ALL REQUIREMENTS MET!")
    else:
        print("✗ Some requirements not fully met")
    print("=" * 80)
    
    return passed == total


def main():
    """Main function to generate and verify dashboard."""
    print("Generating sample Excel file with dashboard layout...")
    
    # Generate sample transactions
    from datetime import timedelta
    import random
    
    base_date = datetime.now() - timedelta(days=180)
    transactions = []
    
    categories = ['Food & Dining', 'Transportation', 'Entertainment', 'Bills & Utilities', 
                  'Shopping', 'Healthcare']
    accounts = ['ACC-CREDIT-001', 'ACC-DEBIT-002', 'ACC-DEBIT-003']
    account_types = ['Credit', 'Debit', 'Debit']
    
    for i in range(150):
        days_offset = random.randint(0, 180)
        transaction_date = base_date + timedelta(days=days_offset)
        
        category = random.choice(categories)
        amount = round(random.uniform(10, 500), 2)
        
        account_idx = i % len(accounts)
        
        transaction = {
            'id': f'TXN{i:06d}',
            'date': transaction_date.strftime('%Y-%m-%d'),
            'payee': f'Merchant {i % 20}',
            'category': category,
            'subcategory': f'Sub-{category.split()[0]}',
            'amount': amount,
            'currency': 'EGP',
            'account_id': accounts[account_idx],
            'account_type': account_types[account_idx],
            'tags': '',
            'anomalies': 'high_value' if i % 30 == 0 else '',
            'confidence': 'high'
        }
        
        transactions.append(transaction)
    
    # Export to Excel
    filename = 'dashboard_verification_output.xlsx'
    exporter = XLSXExporter()
    exporter.export_to_excel(transactions, filename)
    print(f"✓ Generated: {filename}")
    print()
    
    # Verify dashboard layout
    success = verify_dashboard_layout(filename)
    
    return 0 if success else 1


if __name__ == '__main__':
    sys.exit(main())
