"""Unit tests for XLSXExporter."""
import unittest
import os
import tempfile
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import load_workbook
from goldminer.etl import XLSXExporter


class TestXLSXExporter(unittest.TestCase):
    """Test cases for XLSXExporter class."""
    
    def setUp(self):
        """Set up test fixtures."""
        self.exporter = XLSXExporter()
        self.temp_dir = tempfile.mkdtemp()
        
        # Create sample transaction data
        self.sample_transactions = self._create_sample_transactions()
    
    def tearDown(self):
        """Clean up test files."""
        # Remove temporary files
        for file in Path(self.temp_dir).glob("*.xlsx"):
            file.unlink()
        Path(self.temp_dir).rmdir()
    
    def _create_sample_transactions(self):
        """Create sample transaction data for testing."""
        base_date = datetime.now() - timedelta(days=90)
        transactions = []
        
        categories = ['Food', 'Transport', 'Entertainment', 'Bills', 'Shopping']
        accounts = ['ACC001', 'ACC002', 'ACC003']
        
        for i in range(100):
            date = base_date + timedelta(days=i % 30)
            transaction = {
                'id': f'TXN{i:05d}',
                'date': date.strftime('%Y-%m-%d'),
                'payee': f'Merchant {i % 10}',
                'category': categories[i % len(categories)],
                'subcategory': f'Sub{i % 3}',
                'amount': round(50 + (i * 13.7) % 500, 2),
                'currency': 'USD',
                'account_id': accounts[i % len(accounts)],
                'account_type': 'Credit' if i % 2 == 0 else 'Debit',
                'tags': f'tag{i % 5}' if i % 3 == 0 else '',
                'anomalies': 'high_value' if i % 20 == 0 else '',
                'confidence': 'high'
            }
            transactions.append(transaction)
        
        return transactions
    
    def test_initialization(self):
        """Test XLSXExporter initialization."""
        self.assertIsNotNone(self.exporter)
        self.assertIsNotNone(self.exporter.logger)
        self.assertIsNotNone(self.exporter.header_font)
        self.assertIsNotNone(self.exporter.header_fill)
    
    def test_export_to_excel_basic(self):
        """Test basic export functionality."""
        output_file = os.path.join(self.temp_dir, 'test_export.xlsx')
        
        # Export transactions
        self.exporter.export_to_excel(self.sample_transactions, output_file)
        
        # Verify file was created
        self.assertTrue(os.path.exists(output_file))
        
        # Load and verify workbook
        wb = load_workbook(output_file)
        self.assertIn('Transactions', wb.sheetnames)
        self.assertIn('Monthly Summary', wb.sheetnames)
        self.assertIn('Anomalies', wb.sheetnames)
    
    def test_export_empty_transactions(self):
        """Test export with empty transaction list."""
        output_file = os.path.join(self.temp_dir, 'test_empty.xlsx')
        
        with self.assertRaises(ValueError):
            self.exporter.export_to_excel([], output_file)
    
    def test_export_filename_without_extension(self):
        """Test export adds .xlsx extension if missing."""
        output_file = os.path.join(self.temp_dir, 'test_no_ext')
        
        self.exporter.export_to_excel(self.sample_transactions, output_file)
        
        # Verify file was created with .xlsx extension
        expected_file = output_file + '.xlsx'
        self.assertTrue(os.path.exists(expected_file))
    
    def test_transactions_sheet_structure(self):
        """Test Transactions sheet has correct structure."""
        output_file = os.path.join(self.temp_dir, 'test_structure.xlsx')
        
        self.exporter.export_to_excel(self.sample_transactions, output_file)
        
        wb = load_workbook(output_file)
        ws = wb['Transactions']
        
        # Check that headers are present
        self.assertIsNotNone(ws['A1'].value)
        
        # Check that data rows exist
        self.assertGreater(ws.max_row, 1)
        
        # Check that freeze panes is set
        self.assertIsNotNone(ws.freeze_panes)
    
    def test_transactions_sheet_formatting(self):
        """Test Transactions sheet has proper formatting."""
        output_file = os.path.join(self.temp_dir, 'test_formatting.xlsx')
        
        self.exporter.export_to_excel(self.sample_transactions, output_file)
        
        wb = load_workbook(output_file)
        ws = wb['Transactions']
        
        # Check header formatting
        header_cell = ws['A1']
        self.assertTrue(header_cell.font.bold)
        self.assertIsNotNone(header_cell.fill.start_color.rgb)
    
    def test_monthly_summary_sheet(self):
        """Test Monthly Summary sheet is created and has data."""
        output_file = os.path.join(self.temp_dir, 'test_summary.xlsx')
        
        self.exporter.export_to_excel(self.sample_transactions, output_file)
        
        wb = load_workbook(output_file)
        ws = wb['Monthly Summary']
        
        # Check that sheet has content
        self.assertIsNotNone(ws['A1'].value)
        self.assertGreater(ws.max_row, 1)
    
    def test_anomalies_sheet(self):
        """Test Anomalies sheet contains only flagged transactions."""
        output_file = os.path.join(self.temp_dir, 'test_anomalies.xlsx')
        
        self.exporter.export_to_excel(self.sample_transactions, output_file)
        
        wb = load_workbook(output_file)
        ws = wb['Anomalies']
        
        # Check that sheet has content (we have 5 anomalies in sample data)
        self.assertGreater(ws.max_row, 1)
    
    def test_anomalies_sheet_highlighting(self):
        """Test Anomalies sheet has highlighted rows."""
        output_file = os.path.join(self.temp_dir, 'test_anomaly_highlight.xlsx')
        
        self.exporter.export_to_excel(self.sample_transactions, output_file)
        
        wb = load_workbook(output_file)
        ws = wb['Anomalies']
        
        # Check that data rows have fill color (anomaly highlighting)
        if ws.max_row > 1:  # If there are anomalies
            data_cell = ws['A2']  # First data row
            self.assertIsNotNone(data_cell.fill.start_color.rgb)
    
    def test_export_with_missing_columns(self):
        """Test export handles missing columns gracefully."""
        # Create transactions with minimal fields
        minimal_transactions = [
            {
                'id': 'TXN001',
                'date': '2024-01-01',
                'amount': 100.00
            },
            {
                'id': 'TXN002',
                'date': '2024-01-02',
                'amount': 200.00
            }
        ]
        
        output_file = os.path.join(self.temp_dir, 'test_minimal.xlsx')
        
        # Should not raise an error
        self.exporter.export_to_excel(minimal_transactions, output_file)
        
        self.assertTrue(os.path.exists(output_file))
    
    def test_export_with_no_anomalies(self):
        """Test export when no anomalies are present."""
        # Create transactions without anomalies
        clean_transactions = [
            {
                'id': f'TXN{i:03d}',
                'date': '2024-01-01',
                'amount': 100.00,
                'category': 'Food',
                'anomalies': ''
            }
            for i in range(10)
        ]
        
        output_file = os.path.join(self.temp_dir, 'test_no_anomalies.xlsx')
        
        self.exporter.export_to_excel(clean_transactions, output_file)
        
        wb = load_workbook(output_file)
        ws = wb['Anomalies']
        
        # Check that message is displayed
        self.assertIsNotNone(ws['A1'].value)
    
    def test_currency_formatting(self):
        """Test that amount columns have currency formatting."""
        output_file = os.path.join(self.temp_dir, 'test_currency.xlsx')
        
        self.exporter.export_to_excel(self.sample_transactions, output_file)
        
        wb = load_workbook(output_file)
        ws = wb['Transactions']
        
        # Find amount column (assuming it's present)
        headers = [cell.value for cell in ws[1]]
        if 'amount' in headers:
            amount_col_idx = headers.index('amount') + 1
            # Check formatting of a data cell in amount column
            amount_cell = ws.cell(row=2, column=amount_col_idx)
            self.assertIn('$', amount_cell.number_format or '')
    
    def test_charts_in_summary_sheet(self):
        """Test that charts are added to Monthly Summary sheet."""
        output_file = os.path.join(self.temp_dir, 'test_charts.xlsx')
        
        self.exporter.export_to_excel(self.sample_transactions, output_file)
        
        wb = load_workbook(output_file)
        ws = wb['Monthly Summary']
        
        # Check that charts were added
        self.assertGreater(len(ws._charts), 0)
    
    def test_column_width_adjustment(self):
        """Test that column widths are auto-adjusted."""
        output_file = os.path.join(self.temp_dir, 'test_width.xlsx')
        
        self.exporter.export_to_excel(self.sample_transactions, output_file)
        
        wb = load_workbook(output_file)
        ws = wb['Transactions']
        
        # Check that column widths are set
        for column in ws.column_dimensions.values():
            self.assertGreater(column.width, 0)
    
    def test_utf8_compatibility(self):
        """Test export handles UTF-8 characters correctly."""
        utf8_transactions = [
            {
                'id': 'TXN001',
                'date': '2024-01-01',
                'payee': 'Café René',
                'category': 'Food & Drink',
                'amount': 100.00,
                'currency': '€'
            },
            {
                'id': 'TXN002',
                'date': '2024-01-02',
                'payee': '北京饭店',
                'category': 'Restaurant',
                'amount': 200.00,
                'currency': '¥'
            }
        ]
        
        output_file = os.path.join(self.temp_dir, 'test_utf8.xlsx')
        
        # Should handle UTF-8 characters without error
        self.exporter.export_to_excel(utf8_transactions, output_file)
        
        self.assertTrue(os.path.exists(output_file))
        
        # Verify UTF-8 characters are preserved
        wb = load_workbook(output_file)
        ws = wb['Transactions']
        self.assertIsNotNone(ws['A2'].value)
    
    def test_large_dataset_export(self):
        """Test export with a larger dataset."""
        # Create 1000 transactions
        large_transactions = [
            {
                'id': f'TXN{i:05d}',
                'date': (datetime.now() - timedelta(days=i % 365)).strftime('%Y-%m-%d'),
                'payee': f'Merchant {i % 50}',
                'category': f'Category {i % 10}',
                'amount': round(10 + (i * 7.3) % 1000, 2),
                'anomalies': 'high_value' if i % 100 == 0 else ''
            }
            for i in range(1000)
        ]
        
        output_file = os.path.join(self.temp_dir, 'test_large.xlsx')
        
        self.exporter.export_to_excel(large_transactions, output_file)
        
        self.assertTrue(os.path.exists(output_file))
        
        # Verify all sheets are created
        wb = load_workbook(output_file)
        self.assertEqual(len(wb.sheetnames), 3)
    
    def test_header_styling_dark_gray(self):
        """Test headers use dark gray background (professional theme)."""
        output_file = os.path.join(self.temp_dir, 'test_header_color.xlsx')
        
        self.exporter.export_to_excel(self.sample_transactions, output_file)
        
        wb = load_workbook(output_file)
        ws = wb['Transactions']
        
        # Check header cell background color
        header_cell = ws['A1']
        fill_color = header_cell.fill.start_color.rgb
        
        # Should be dark gray (#404040), not blue
        self.assertEqual(fill_color, '00404040')
        self.assertNotEqual(fill_color, '00366092')  # Old blue color
    
    def test_calibri_font_consistency(self):
        """Test all cells use Calibri font."""
        output_file = os.path.join(self.temp_dir, 'test_calibri_font.xlsx')
        
        self.exporter.export_to_excel(self.sample_transactions, output_file)
        
        wb = load_workbook(output_file)
        
        for sheet_name in ['Transactions', 'Anomalies']:
            ws = wb[sheet_name]
            
            # Check header font
            header_cell = ws['A1']
            self.assertEqual(header_cell.font.name, 'Calibri')
            self.assertEqual(header_cell.font.size, 11)
            
            # Check data cell font (if exists)
            if ws.max_row > 1:
                data_cell = ws['A2']
                if data_cell.value is not None:
                    self.assertEqual(data_cell.font.name, 'Calibri')
                    self.assertEqual(data_cell.font.size, 11)
    
    def test_alternating_row_shading(self):
        """Test alternating row shading is applied to data tables."""
        output_file = os.path.join(self.temp_dir, 'test_alternating_rows.xlsx')
        
        self.exporter.export_to_excel(self.sample_transactions, output_file)
        
        wb = load_workbook(output_file)
        ws = wb['Transactions']
        
        # Check that some rows have shading
        has_shading = False
        for row_idx in range(2, min(10, ws.max_row + 1)):
            cell = ws.cell(row=row_idx, column=1)
            if cell.fill.start_color.rgb and cell.fill.start_color.rgb not in ['00000000', 'FFFFFFFF']:
                has_shading = True
                break
        
        self.assertTrue(has_shading, "No alternating row shading detected")
    
    def test_alternating_row_pattern(self):
        """Test that alternating row shading follows odd/even pattern."""
        output_file = os.path.join(self.temp_dir, 'test_alternating_pattern.xlsx')
        
        self.exporter.export_to_excel(self.sample_transactions, output_file)
        
        wb = load_workbook(output_file)
        ws = wb['Transactions']
        
        # Check pattern: row 2 should be white, row 3 should be shaded, row 4 white, etc.
        row2_fill = ws.cell(row=2, column=1).fill.start_color.rgb
        row3_fill = ws.cell(row=3, column=1).fill.start_color.rgb
        
        # The pattern should be different (one shaded, one not)
        self.assertNotEqual(row2_fill, row3_fill)
    
    def test_reusable_style_templates(self):
        """Test that style templates are properly defined and reusable."""
        # Check that exporter has all required style templates
        self.assertIsNotNone(self.exporter.default_font)
        self.assertIsNotNone(self.exporter.header_font)
        self.assertIsNotNone(self.exporter.header_fill)
        self.assertIsNotNone(self.exporter.alternating_fill)
        self.assertIsNotNone(self.exporter.currency_font)
        self.assertIsNotNone(self.exporter.tag_font)
        
        # Verify font names
        self.assertEqual(self.exporter.default_font.name, 'Calibri')
        self.assertEqual(self.exporter.header_font.name, 'Calibri')
        self.assertEqual(self.exporter.currency_font.name, 'Calibri')
        self.assertEqual(self.exporter.tag_font.name, 'Calibri')
        
        # Verify header styling
        self.assertEqual(self.exporter.header_fill.start_color.rgb, '00404040')
        self.assertTrue(self.exporter.header_font.bold)
    
    def test_consistency_across_sheets(self):
        """Test styling consistency across all three sheets."""
        output_file = os.path.join(self.temp_dir, 'test_consistency.xlsx')
        
        self.exporter.export_to_excel(self.sample_transactions, output_file)
        
        wb = load_workbook(output_file)
        
        # Check each sheet has consistent styling
        for sheet_name in ['Transactions', 'Anomalies']:
            ws = wb[sheet_name]
            
            # Check freeze panes
            self.assertIsNotNone(ws.freeze_panes)
            
            # Check header styling
            header_cell = ws['A1']
            self.assertTrue(header_cell.font.bold)
            self.assertEqual(header_cell.font.name, 'Calibri')
            
            # Check column widths are adjusted
            adjusted_widths = sum(1 for col in ws.column_dimensions.values() 
                                 if col.width and col.width > 8.43)
            self.assertGreater(adjusted_widths, 0)
    
    def test_dashboard_layout(self):
        """Test dashboard-style layout in Monthly Summary sheet."""
        output_file = os.path.join(self.temp_dir, 'test_dashboard.xlsx')
        
        self.exporter.export_to_excel(self.sample_transactions, output_file)
        
        wb = load_workbook(output_file)
        ws = wb['Monthly Summary']
        
        # Check dashboard header
        self.assertEqual(ws['A1'].value, 'GoldMiner Financial Dashboard')
        self.assertTrue(ws['A1'].font.bold)
        # Check that A1 is merged across multiple columns
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        self.assertTrue(any('A1' in r and 'N1' in r for r in merged_ranges))
        
        # Check Key Metrics section
        self.assertEqual(ws['A3'].value, 'Key Metrics')
        self.assertIn('Total Spend', str(ws['A5'].value))
        self.assertIn('# of Transactions', str(ws['A7'].value))
        self.assertIn('Credit Card', str(ws['A9'].value))
        self.assertIn('Debit Card', str(ws['A10'].value))
        
        # Check that metric values are numbers
        self.assertIsNotNone(ws['B5'].value)
        self.assertIsInstance(ws['B5'].value, (int, float))
        self.assertIsNotNone(ws['B7'].value)
        self.assertIsInstance(ws['B7'].value, (int, float))
        
        # Check number formatting includes currency symbol
        self.assertIn('#,##0.00', ws['B5'].number_format)
        # Currency symbol should be present (USD, EGP, etc.)
        self.assertTrue(any(c in ws['B5'].number_format for c in ['USD', 'EGP', '$']))
        
        # Check pivot table section
        self.assertEqual(ws['A22'].value, 'Monthly Category Breakdown')
        self.assertEqual(ws['A24'].value, 'category')  # Pivot table header
        
        # Check freeze panes at A4
        self.assertEqual(ws.freeze_panes, 'A4')
        
        # Check page setup
        self.assertEqual(ws.page_setup.orientation, ws.ORIENTATION_LANDSCAPE)
        self.assertEqual(ws.page_setup.fitToWidth, 1)
        
        # Check footer
        self.assertIn('GoldMiner Report', ws.oddFooter.center.text)
        self.assertIn('Generated on', ws.oddFooter.center.text)
        
        # Check charts are present
        self.assertEqual(len(ws._charts), 3)
        
        # Verify chart titles
        chart_titles = []
        for chart in ws._charts:
            if chart.title and chart.title.tx and chart.title.tx.rich:
                for para in chart.title.tx.rich.p:
                    for run in para.r:
                        if run.t:
                            chart_titles.append(run.t)
        
        self.assertIn('Spending by Category', chart_titles)
        self.assertIn('Monthly Spending', chart_titles)
        self.assertIn('Cumulative Spending', chart_titles)


if __name__ == '__main__':
    unittest.main()
