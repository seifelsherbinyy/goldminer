"""Unit tests for XLSXExporter conditional formatting features."""
import unittest
import os
import tempfile
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import load_workbook
from goldminer.etl import XLSXExporter


class TestXLSXConditionalFormatting(unittest.TestCase):
    """Test cases for XLSXExporter conditional formatting."""
    
    def setUp(self):
        """Set up test fixtures."""
        self.exporter = XLSXExporter()
        self.temp_dir = tempfile.mkdtemp()
    
    def tearDown(self):
        """Clean up test files."""
        # Remove temporary files
        for file in Path(self.temp_dir).glob("*.xlsx"):
            file.unlink()
        Path(self.temp_dir).rmdir()
    
    def _create_transactions_with_urgency(self, num_transactions=100):
        """Create sample transaction data with urgency field."""
        base_date = datetime.now() - timedelta(days=90)
        transactions = []
        
        categories = ['Food', 'Transport', 'Entertainment', 'Bills', 'Shopping']
        accounts = ['ACC001', 'ACC002', 'ACC003']
        urgency_levels = ['high', 'medium', 'normal']
        anomaly_types = ['high_value', 'burst_frequency', 'unknown_merchant', '']
        
        for i in range(num_transactions):
            date = base_date + timedelta(days=i % 30)
            
            # Vary urgency levels
            if i % 10 == 0:
                urgency = 'high'
                amount = 15000 + (i * 100)
            elif i % 5 == 0:
                urgency = 'medium'
                amount = 7500 + (i * 50)
            else:
                urgency = 'normal'
                amount = 50 + (i * 13.7) % 500
            
            # Add various anomaly types
            anomaly_flag = anomaly_types[i % len(anomaly_types)]
            
            transaction = {
                'id': f'TXN{i:05d}',
                'date': date.strftime('%Y-%m-%d'),
                'payee': f'Merchant {i % 10}',
                'category': categories[i % len(categories)],
                'subcategory': f'Sub{i % 3}',
                'amount': round(amount, 2),
                'currency': 'USD',
                'account_id': accounts[i % len(accounts)],
                'account_type': 'Credit' if i % 2 == 0 else 'Debit',
                'urgency': urgency,
                'tags': f'tag{i % 5}' if i % 3 == 0 else '',
                'anomalies': anomaly_flag,
                'confidence': 'high'
            }
            transactions.append(transaction)
        
        return transactions
    
    def test_urgency_field_included_in_transactions_sheet(self):
        """Test that urgency field is included in Transactions sheet."""
        transactions = self._create_transactions_with_urgency(50)
        output_file = os.path.join(self.temp_dir, 'test_urgency_field.xlsx')
        
        self.exporter.export_to_excel(transactions, output_file)
        
        wb = load_workbook(output_file)
        ws = wb['Transactions']
        
        # Check that urgency column exists in headers
        headers = [cell.value for cell in ws[1]]
        self.assertIn('urgency', headers)
        
        # Check that urgency values are present
        urgency_col_idx = headers.index('urgency') + 1
        urgency_values = [ws.cell(row=i, column=urgency_col_idx).value 
                         for i in range(2, min(10, ws.max_row + 1))]
        
        # Verify we have urgency values
        self.assertTrue(any(v in ['high', 'medium', 'normal'] for v in urgency_values))
    
    def test_urgency_field_included_in_anomalies_sheet(self):
        """Test that urgency field is included in Anomalies sheet."""
        transactions = self._create_transactions_with_urgency(50)
        output_file = os.path.join(self.temp_dir, 'test_urgency_anomalies.xlsx')
        
        self.exporter.export_to_excel(transactions, output_file)
        
        wb = load_workbook(output_file)
        ws = wb['Anomalies']
        
        # Check that urgency column exists in headers
        headers = [cell.value for cell in ws[1]]
        self.assertIn('urgency', headers)
    
    def test_conditional_formatting_rules_applied_to_transactions(self):
        """Test that conditional formatting rules are applied to Transactions sheet."""
        transactions = self._create_transactions_with_urgency(50)
        output_file = os.path.join(self.temp_dir, 'test_cf_transactions.xlsx')
        
        self.exporter.export_to_excel(transactions, output_file)
        
        wb = load_workbook(output_file)
        ws = wb['Transactions']
        
        # Check that conditional formatting rules exist
        self.assertGreater(len(ws.conditional_formatting._cf_rules), 0)
    
    def test_conditional_formatting_rules_applied_to_anomalies(self):
        """Test that conditional formatting rules are applied to Anomalies sheet."""
        transactions = self._create_transactions_with_urgency(50)
        output_file = os.path.join(self.temp_dir, 'test_cf_anomalies.xlsx')
        
        self.exporter.export_to_excel(transactions, output_file)
        
        wb = load_workbook(output_file)
        ws = wb['Anomalies']
        
        # Check that conditional formatting rules exist if anomalies present
        if ws.max_row > 1:  # More than just header
            self.assertGreater(len(ws.conditional_formatting._cf_rules), 0)
    
    def test_anomaly_borders_applied(self):
        """Test that anomaly borders are applied to flagged rows."""
        transactions = self._create_transactions_with_urgency(50)
        output_file = os.path.join(self.temp_dir, 'test_anomaly_borders.xlsx')
        
        self.exporter.export_to_excel(transactions, output_file)
        
        wb = load_workbook(output_file)
        ws = wb['Transactions']
        
        # Find a row with anomalies
        headers = [cell.value for cell in ws[1]]
        anomaly_col_idx = headers.index('anomalies') + 1
        
        for row_idx in range(2, min(ws.max_row + 1, 20)):
            anomaly_value = ws.cell(row=row_idx, column=anomaly_col_idx).value
            if anomaly_value and str(anomaly_value).strip():
                # Check that cells in this row have the anomaly border
                cell = ws.cell(row=row_idx, column=1)
                self.assertIsNotNone(cell.border)
                # Verify border style is thick (anomaly border)
                self.assertEqual(cell.border.left.style, 'thick')
                break
    
    def test_large_dataset_with_conditional_formatting(self):
        """Test conditional formatting with 1000+ records."""
        # Create 1200 transactions with mixed urgency and anomalies
        transactions = self._create_transactions_with_urgency(1200)
        output_file = os.path.join(self.temp_dir, 'test_large_cf.xlsx')
        
        self.exporter.export_to_excel(transactions, output_file)
        
        # Verify file was created
        self.assertTrue(os.path.exists(output_file))
        
        # Load and verify
        wb = load_workbook(output_file)
        ws = wb['Transactions']
        
        # Verify all data is present
        self.assertEqual(ws.max_row, 1201)  # 1200 data rows + 1 header
        
        # Verify conditional formatting rules are applied
        self.assertGreater(len(ws.conditional_formatting._cf_rules), 0)
        
        # Verify urgency column exists
        headers = [cell.value for cell in ws[1]]
        self.assertIn('urgency', headers)
        
        # Count different urgency levels in data
        urgency_col_idx = headers.index('urgency') + 1
        urgency_counts = {'high': 0, 'medium': 0, 'normal': 0}
        
        for row_idx in range(2, min(100, ws.max_row + 1)):  # Sample first 100 rows
            urgency_value = ws.cell(row=row_idx, column=urgency_col_idx).value
            if urgency_value in urgency_counts:
                urgency_counts[urgency_value] += 1
        
        # Verify we have a mix of urgency levels
        self.assertGreater(urgency_counts['high'], 0)
        self.assertGreater(urgency_counts['medium'], 0)
        self.assertGreater(urgency_counts['normal'], 0)
    
    def test_anomalies_sheet_formatting_with_urgency(self):
        """Test Anomalies sheet has both urgency and anomaly formatting."""
        transactions = self._create_transactions_with_urgency(100)
        output_file = os.path.join(self.temp_dir, 'test_anomalies_urgency.xlsx')
        
        self.exporter.export_to_excel(transactions, output_file)
        
        wb = load_workbook(output_file)
        ws = wb['Anomalies']
        
        if ws.max_row > 1:  # If there are anomalies
            headers = [cell.value for cell in ws[1]]
            
            # Verify both urgency and anomalies columns exist
            self.assertIn('urgency', headers)
            self.assertIn('anomalies', headers)
            
            # Verify conditional formatting rules are applied
            self.assertGreater(len(ws.conditional_formatting._cf_rules), 0)
            
            # Verify at least one row has anomaly borders
            anomaly_col_idx = headers.index('anomalies') + 1
            found_border = False
            
            for row_idx in range(2, min(ws.max_row + 1, 10)):
                anomaly_value = ws.cell(row=row_idx, column=anomaly_col_idx).value
                if anomaly_value and str(anomaly_value).strip():
                    cell = ws.cell(row=row_idx, column=1)
                    if cell.border and cell.border.left.style == 'thick':
                        found_border = True
                        break
            
            self.assertTrue(found_border, "Expected to find at least one row with anomaly border")
    
    def test_mixed_urgency_types_in_large_dataset(self):
        """Test that all urgency types are present and correctly formatted in a large dataset."""
        transactions = self._create_transactions_with_urgency(1500)
        output_file = os.path.join(self.temp_dir, 'test_mixed_urgency.xlsx')
        
        self.exporter.export_to_excel(transactions, output_file)
        
        wb = load_workbook(output_file)
        ws = wb['Transactions']
        
        headers = [cell.value for cell in ws[1]]
        urgency_col_idx = headers.index('urgency') + 1
        anomaly_col_idx = headers.index('anomalies') + 1
        
        # Count urgency types and anomalies
        urgency_counts = {'high': 0, 'medium': 0, 'normal': 0}
        anomaly_count = 0
        
        for row_idx in range(2, ws.max_row + 1):
            urgency_value = ws.cell(row=row_idx, column=urgency_col_idx).value
            if urgency_value in urgency_counts:
                urgency_counts[urgency_value] += 1
            
            anomaly_value = ws.cell(row=row_idx, column=anomaly_col_idx).value
            if anomaly_value and str(anomaly_value).strip():
                anomaly_count += 1
        
        # Verify distribution
        total = sum(urgency_counts.values())
        self.assertEqual(total, 1500)
        
        # Based on our generation logic:
        # - Every 10th transaction is high (150 transactions)
        # - Every 5th (but not 10th) is medium (150 transactions)
        # - Rest are normal (1200 transactions)
        self.assertEqual(urgency_counts['high'], 150)
        self.assertEqual(urgency_counts['medium'], 150)
        self.assertEqual(urgency_counts['normal'], 1200)
        
        # Verify anomalies exist
        self.assertGreater(anomaly_count, 0)
    
    def test_helper_function_apply_urgency_formatting(self):
        """Test the helper function for applying urgency formatting."""
        transactions = self._create_transactions_with_urgency(50)
        output_file = os.path.join(self.temp_dir, 'test_helper_urgency.xlsx')
        
        self.exporter.export_to_excel(transactions, output_file)
        
        # Just verify the export succeeded and file has content
        wb = load_workbook(output_file)
        ws = wb['Transactions']
        
        # Verify conditional formatting was applied via helper
        self.assertGreater(len(ws.conditional_formatting._cf_rules), 0)
    
    def test_helper_function_apply_anomaly_borders(self):
        """Test the helper function for applying anomaly borders."""
        transactions = self._create_transactions_with_urgency(50)
        output_file = os.path.join(self.temp_dir, 'test_helper_borders.xlsx')
        
        self.exporter.export_to_excel(transactions, output_file)
        
        wb = load_workbook(output_file)
        ws = wb['Transactions']
        
        headers = [cell.value for cell in ws[1]]
        anomaly_col_idx = headers.index('anomalies') + 1
        
        # Find and verify at least one anomaly row has borders
        found_anomaly_with_border = False
        for row_idx in range(2, min(ws.max_row + 1, 30)):
            anomaly_value = ws.cell(row=row_idx, column=anomaly_col_idx).value
            if anomaly_value and str(anomaly_value).strip():
                cell = ws.cell(row=row_idx, column=1)
                if cell.border and cell.border.left.style == 'thick':
                    found_anomaly_with_border = True
                    break
        
        self.assertTrue(found_anomaly_with_border)
    
    def test_no_urgency_field_graceful_handling(self):
        """Test that export works gracefully when urgency field is missing."""
        # Create transactions without urgency field
        transactions = [
            {
                'id': f'TXN{i:03d}',
                'date': '2024-01-01',
                'amount': 100.00,
                'category': 'Food',
                'anomalies': 'high_value' if i % 10 == 0 else ''
            }
            for i in range(50)
        ]
        
        output_file = os.path.join(self.temp_dir, 'test_no_urgency.xlsx')
        
        # Should not raise an error
        self.exporter.export_to_excel(transactions, output_file)
        
        # Verify file was created
        self.assertTrue(os.path.exists(output_file))
        
        wb = load_workbook(output_file)
        ws = wb['Transactions']
        
        # Verify urgency column is not present
        headers = [cell.value for cell in ws[1]]
        self.assertNotIn('urgency', headers)


if __name__ == '__main__':
    unittest.main()
